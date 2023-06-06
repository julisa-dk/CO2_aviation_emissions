import openpyxl
import csv
from openpyxl import load_workbook
import os

import boto3
from botocore.exceptions import NoCredentialsError


os.system("wget https://www.eurocontrol.int/performance/data/download/xls/CO2_emissions_by_state.xlsx")

inputExcelFile = 'CO2_emissions_by_state.xlsx'

newWorkbook = openpyxl.load_workbook(inputExcelFile)

print("Found the following worksheets:")
for sheetname in newWorkbook.sheetnames:
    print(sheetname)

firstWorksheet = newWorkbook['DATA']
all_rows = list(firstWorksheet.rows)
print(f"Found {len(all_rows)} rows of data.")

OutputCsvFile = csv.writer(open("ResultCsvFile.csv", 'w'), delimiter=",")

for eachrow in firstWorksheet.rows:
    OutputCsvFile.writerow([cell.value for cell in eachrow])

# Print out bucket names
""" s3 = boto3.resource("s3")
for bucket in s3.buckets.all():
    print(bucket.name) """

bucket = 's3emission'
local_file = 'ResultCsvFile.csv'

s3 = boto3.client("s3")
try:
    with open(local_file, "rb") as f:
        s3.upload_fileobj(f, bucket, local_file)
        print("The file upload successfully")
except Exception as e:
        print(f"Error uploading file to S3: {e}")
