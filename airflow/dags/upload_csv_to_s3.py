from datetime import datetime, timedelta
import os

import openpyxl
import csv
from openpyxl import load_workbook

from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.amazon.aws.hooks.s3 import S3Hook

from airflow.operators.empty import EmptyOperator
from airflow.operators.bash import BashOperator

CWD = os.getcwd()
PATH_DATA = os.path.abspath(os.path.join(CWD, 'dags'))
URL = 'https://www.eurocontrol.int/performance/data/download/xls/CO2_emissions_by_state.xlsx'

def web_data_transform():
    inputExcelFile = f"{PATH_DATA}/CO2_emissions_by_state.xlsx"

    newWorkbook = openpyxl.load_workbook(inputExcelFile)

    print("Found the following worksheets:")
    for sheetname in newWorkbook.sheetnames:
      print(sheetname)

    firstWorksheet = newWorkbook['DATA']
    all_rows = list(firstWorksheet.rows)
    print(f"Found {len(all_rows)} rows of data.")

    OutputCsvFile = csv.writer(open(f"{PATH_DATA}/ResultCsvFile.csv", 'w'), delimiter=",")

    for eachrow in firstWorksheet.rows:
      OutputCsvFile.writerow([cell.value for cell in eachrow])

def upload_to_s3(filename: str, key: str, bucket_name: str) -> None:
    hook = S3Hook('s3_conn')
    hook.load_file(filename=filename, key=key, bucket_name=bucket_name)

# initializing the default arguments
default_args = {
		'owner': 'Julia',
		'start_date': datetime(2023, 5, 20),
		'retries': 1,
		'retry_delay': timedelta(minutes=5)
}

with DAG(
    dag_id='s3_dag',
    default_args=default_args,
    schedule_interval='@daily',
    catchup=False
) as dag:
    
    start_task = EmptyOperator(
        task_id='start'
    )

    loading_file = BashOperator(
        task_id = 'loading_file',
        bash_command = 'curl https://www.eurocontrol.int/performance/data/download/xls/CO2_emissions_by_state.xlsx -o /opt/airflow/dags/CO2_emissions_by_state.xlsx'
    )

    data_transform = PythonOperator(
        task_id='web_data',
        python_callable=web_data_transform
    )

    # Upload the file
    task_upload_to_s3 = PythonOperator(
        task_id='upload_to_s3',
        python_callable=upload_to_s3,
        op_kwargs={
            'filename': f"{PATH_DATA}/ResultCsvFile.csv",
            'key': 'CsvFile.csv',
            'bucket_name': 's3emission'
        }
        
    )

    end_task = EmptyOperator(
        task_id='end'
    )

start_task >> loading_file
loading_file >> data_transform
data_transform >> task_upload_to_s3
task_upload_to_s3 >> end_task
